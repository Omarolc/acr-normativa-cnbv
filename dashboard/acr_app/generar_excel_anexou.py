"""
Generador del reporte regulatorio Anexo U en Excel
SOCAP con Nivel de Operaciones Básico — 3 hojas:
  1. Balance General
  2. Estado de Resultados
  3. Cómputo del Nivel de Capitalización
"""
import io
import openpyxl
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, numbers
)
from openpyxl.utils import get_column_letter

# ── Paleta ────────────────────────────────────────────────────────────────────
NEGRO       = "00000000"
GRIS_HDR    = "00D9D9D9"
GRIS_FILA   = "00F2F2F2"
VERDE_A     = "00C6EFCE"; VERDE_TXT   = "00276221"
AMARILLO_B  = "00FFEB9C"; AMARILLO_TXT= "009C5700"
NARANJA_C   = "00FFCC99"; NARANJA_TXT = "00974706"
ROJO_D      = "00FFC7CE"; ROJO_TXT    = "009C0006"
AZUL_TITULO = "001F3864"
AZUL_SUB    = "002F5597"
BLANCO      = "00FFFFFF"

FMT_PESO    = '"$"#,##0.00_);("$"#,##0.00)'
FMT_PESO_N  = '"$"#,##0.00;("$"#,##0.00)'
FMT_PCT     = '0.00"%"'

def thin():
    s = Side(style="thin", color=NEGRO)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_only(color=NEGRO):
    return Border(bottom=Side(style="thin", color=color))

def thick_bottom():
    return Border(bottom=Side(style="medium", color=NEGRO))

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color=NEGRO, italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def alin(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def cat_colors(cat):
    return {
        "A": (VERDE_A, VERDE_TXT),
        "B": (AMARILLO_B, AMARILLO_TXT),
        "C": (NARANJA_C, NARANJA_TXT),
        "D": (ROJO_D, ROJO_TXT),
    }.get(cat, (GRIS_HDR, NEGRO))

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ── Helpers de escritura ──────────────────────────────────────────────────────
def write(ws, row, col, value, bold=False, size=10, color=NEGRO,
          h="left", fill_hex=None, border=None, fmt=None, italic=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = font(bold=bold, size=size, color=color, italic=italic)
    cell.alignment = alin(h=h)
    if fill_hex:   cell.fill   = fill(fill_hex)
    if border:     cell.border = border
    if fmt:        cell.number_format = fmt
    return cell

def merge_write(ws, r1, c1, r2, c2, value, bold=False, size=10,
                color=NEGRO, h="center", fill_hex=None):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    cell.font      = font(bold=bold, size=size, color=color)
    cell.alignment = alin(h=h, v="center")
    if fill_hex: cell.fill = fill(fill_hex)
    return cell

# ── Encabezado común ──────────────────────────────────────────────────────────
def encabezado(ws, nombre, titulo, subtitulo, col_fin=6):
    merge_write(ws, 1, 1, 1, col_fin, nombre,
                bold=True, size=12, color=BLANCO, fill_hex=AZUL_TITULO)
    merge_write(ws, 2, 1, 2, col_fin, titulo,
                bold=True, size=11, color=BLANCO, fill_hex=AZUL_SUB)
    merge_write(ws, 3, 1, 3, col_fin, subtitulo,
                bold=False, size=9, color=NEGRO, h="center")
    merge_write(ws, 4, 1, 4, col_fin, "(Cifras en pesos)",
                bold=False, size=8, color="00666666", h="center", fill_hex=GRIS_FILA)
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 18

# ── HOJA 1: Balance General ───────────────────────────────────────────────────
def hoja_balance(wb, d, c, nombre, fecha_corte):
    ws = wb.create_sheet("Balance General")
    ws.sheet_view.showGridLines = False

    # Anchos de columna: A concepto_activo | B importe_a | C espacio | D concepto_pasivo | E importe_p
    for col, w in [(1,32),(2,16),(3,2),(4,32),(5,16)]:
        set_col_width(ws, col, w)

    encabezado(ws, nombre,
               f"BALANCE GENERAL AL {fecha_corte.upper()}",
               "", col_fin=5)

    # Sub-encabezados
    r = 6
    for col, txt in [(1,"ACTIVO"),(4,"PASIVO Y CAPITAL")]:
        cell = ws.cell(row=r, column=col, value=txt)
        cell.font      = font(bold=True, size=10, color=BLANCO)
        cell.fill      = fill(AZUL_SUB)
        cell.alignment = alin(h="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    r += 1

    def fila_bal(row, tl, vl, tr, vr, bold_l=False, bold_r=False,
                 fill_l=None, fill_r=None, sep_l=False, sep_r=False):
        def w(col, txt, v, bold, fh, sep, fmt=FMT_PESO_N):
            if txt is not None:
                c2 = ws.cell(row=row, column=col, value=txt)
                c2.font      = font(bold=bold, size=9)
                c2.fill      = fill(fh) if fh else fill(BLANCO)
                c2.alignment = alin(h="left")
                brd = thick_bottom() if sep else bottom_only("00CCCCCC")
                c2.border = brd
            if v is not None:
                c3 = ws.cell(row=row, column=col+1, value=v)
                c3.font           = font(bold=bold, size=9)
                c3.number_format  = fmt
                c3.alignment      = alin(h="right")
                c3.fill           = fill(fh) if fh else fill(BLANCO)
                c3.border         = thick_bottom() if sep else bottom_only("00CCCCCC")
        w(1, tl, vl, bold_l, fill_l, sep_l)
        w(4, tr, vr, bold_r, fill_r, sep_r)

    cv = d["cartera_vigente"]
    cvenc = d["cartera_vencida"]
    ep = d["estimacion_prev"]
    cn = c["cn"]

    filas = [
        ("EFECTIVO",                    d["efectivo"],
         "DEPÓSITOS",                   None, True, True, GRIS_FILA, GRIS_FILA),
        (None, None,
         "  Exigibilidad inmediata",    d["dep_exig_inm"], False, False, None, None),
        ("CARTERA DE CRÉDITO VIGENTE",  cv,
         "  A plazo",                   d["dep_plazo"], True, False, None, None),
        ("CARTERA DE CRÉDITO VENCIDA",  cvenc,
         "  Sin movimiento",            d["cuentas_sin_mov"], True, False, None, None),
        ("TOTAL CARTERA DE CRÉDITO",    cv+cvenc,
         None, None, True, False, GRIS_FILA, None),
        ("(−) ESTIMACIÓN PREVENTIVA",   ep,
         "PRÉSTAMOS BANCARIOS", None, True, True, None, GRIS_FILA),
        ("CARTERA DE CRÉDITO (NETO)",   cn,
         "  Corto plazo",               d["prestamos_cp"], True, False, GRIS_FILA, None),
        ("OTRAS CxC (NETO)",            d["otras_cxc"],
         "  Largo plazo",               d["prestamos_lp"], False, False, None, None),
        ("BIENES ADJUDICADOS",          d["bienes_adj"],
         "OTRAS CUENTAS POR PAGAR",     d["otras_cxp"], False, True, None, GRIS_FILA),
        ("INMUEBLES Y EQUIPO (NETO)",   d["inmuebles"],
         "TOTAL PASIVO",                c["tp"], False, True, None, GRIS_FILA),
        ("OTROS ACTIVOS",               d["otros_activos"],
         "CAPITAL CONTRIBUIDO",         None, False, True, None, GRIS_FILA),
        (None, None,
         "  Capital social / Cert. ord.", d["cert_ord"], False, False, None, None),
        (None, None,
         "  Certificados excedentes",   d["cert_vol"], False, False, None, None),
        (None, None,
         "CAPITAL GANADO",              None, False, True, None, GRIS_FILA),
        (None, None,
         "  Reservas de capital",       d["reservas"], False, False, None, None),
        (None, None,
         "  Resultado ejercicios ant.", d["result_ant"], False, False, None, None),
        (None, None,
         "  Resultado neto del periodo",d["result_neto_bg"], False, False, None, None),
        (None, None,
         "TOTAL CAPITAL CONTABLE",      c["cc"], False, True, None, GRIS_FILA),
    ]

    for tl, vl, tr, vr, bl, br, fl, fr in filas:
        fila_bal(r, tl, vl, tr, vr, bl, br, fl, fr)
        ws.row_dimensions[r].height = 15
        r += 1

    # Totales finales
    r += 1
    for col, txt, val in [(1,"TOTAL ACTIVO", c["ta"]),
                          (4,"TOTAL PASIVO Y CAPITAL CONTABLE", c["tp"]+c["cc"])]:
        c1 = ws.cell(row=r, column=col, value=txt)
        c1.font=font(bold=True,size=10,color=BLANCO); c1.fill=fill(AZUL_TITULO)
        c1.alignment=alin(h="left")
        c2 = ws.cell(row=r, column=col+1, value=val)
        c2.font=font(bold=True,size=10,color=BLANCO); c2.fill=fill(AZUL_TITULO)
        c2.number_format=FMT_PESO_N; c2.alignment=alin(h="right")
    ws.row_dimensions[r].height = 18
    r += 2

    # Cuentas de orden
    cell = ws.cell(row=r, column=1, value="CUENTAS DE ORDEN")
    cell.font=font(bold=True,size=9,color=BLANCO); cell.fill=fill(AZUL_SUB)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    for lbl, val in [
        ("Compromisos crediticios",                d.get("compromisos",0) or 0),
        ("Int. devengados no cobrados (vencida)",  d.get("int_dev_nc",0) or 0),
        ("Otras cuentas de registro",              d.get("otras_orden",0) or 0),
    ]:
        ws.cell(row=r,column=1,value=lbl).font=font(size=9)
        c2=ws.cell(row=r,column=2,value=val)
        c2.font=font(size=9); c2.number_format=FMT_PESO_N; c2.alignment=alin(h="right")
        r+=1

    r += 1
    nota = ("La formulación y presentación de los estados financieros básicos es responsabilidad "
            "del Consejo de Administración (Art. 1 Bis 1 Disposiciones CNBV). "
            "Contraparte: Comité de Supervisión Auxiliar (FOCOOP).")
    cell=ws.cell(row=r,column=1,value=nota)
    cell.font=font(size=7.5,italic=True,color="00666666")
    cell.alignment=alin(wrap=True)
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
    ws.row_dimensions[r].height=28

# ── HOJA 2: Estado de Resultados ─────────────────────────────────────────────
def hoja_er(wb, d, c, nombre, fecha_corte):
    ws = wb.create_sheet("Estado de Resultados")
    ws.sheet_view.showGridLines = False
    for col, w in [(1,46),(2,18)]: set_col_width(ws, col, w)

    encabezado(ws, nombre,
               "ESTADO DE RESULTADOS",
               f"DEL 1 DE ENERO AL {fecha_corte.upper()}", col_fin=2)

    r = 6
    def fila_er(concepto, valor, bold=False, sep=False, fill_hex=None):
        nonlocal r
        c1 = ws.cell(row=r, column=1, value=concepto)
        c1.font=font(bold=bold,size=9)
        c1.alignment=alin()
        if fill_hex: c1.fill=fill(fill_hex)
        if sep: c1.border=thick_bottom()

        c2 = ws.cell(row=r, column=2, value=valor)
        c2.font=font(bold=bold,size=9)
        c2.number_format=FMT_PESO_N
        c2.alignment=alin(h="right")
        if fill_hex: c2.fill=fill(fill_hex)
        if sep: c2.border=thick_bottom()
        ws.row_dimensions[r].height=15
        r += 1

    fila_er("Ingresos por intereses",        d["ingresos_int"])
    fila_er("(−) Gastos por intereses",     -d["gastos_int"])
    fila_er("RESULTADO FINANCIERO",          c["rf"],   bold=True, sep=True, fill_hex=GRIS_FILA)
    fila_er("(−) Estimación preventiva",    -d["est_riesgos_er"])
    fila_er("RESULTADO FINANCIERO AJUSTADO POR RIESGOS CREDITICIOS",
                                             c["rfaj"], bold=True, sep=True, fill_hex=GRIS_FILA)
    fila_er("Otros ingresos (egresos) de la operación", d["otros_ing"])
    fila_er("(−) Gastos de administración y promoción", -d["gastos_adm"])
    fila_er("RESULTADO NETO",                c["rn"],   bold=True, sep=True, fill_hex=AZUL_TITULO)
    # color blanco en resultado neto
    for col in [1,2]:
        ws.cell(row=r-1,column=col).font=font(bold=True,size=9,color=BLANCO)
        ws.cell(row=r-1,column=col).fill=fill(AZUL_TITULO)

    r += 2
    nota = ("La formulación y presentación es responsabilidad del Consejo de Administración "
            "(Art. 1 Bis 1). Contraparte: Comité de Supervisión Auxiliar (FOCOOP).")
    cell=ws.cell(row=r,column=1,value=nota)
    cell.font=font(size=7.5,italic=True,color="00666666")
    cell.alignment=alin(wrap=True)
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
    ws.row_dimensions[r].height=28

# ── HOJA 3: Cómputo Nivel de Capitalización ───────────────────────────────────
def hoja_cap(wb, d, c, nombre, fecha_corte):
    ws = wb.create_sheet("Nivel de Capitalización")
    ws.sheet_view.showGridLines = False
    for col, w in [(1,6),(2,40),(3,20),(4,18)]: set_col_width(ws, col, w)

    encabezado(ws, nombre,
               "CÓMPUTO DEL NIVEL DE CAPITALIZACIÓN",
               f"CIFRAS AL {fecha_corte.upper()} · Arts. 1 Bis 3, 1 Bis 4 y 1 Bis 6 Disposiciones CNBV",
               col_fin=4)

    r = 6
    # Encabezado de columnas
    for col, txt, h in [(1,"#","center"),(2,"Concepto","left"),
                        (3,"Fundamento","center"),(4,"Importe","right")]:
        cell=ws.cell(row=r,column=col,value=txt)
        cell.font=font(bold=True,size=9,color=BLANCO)
        cell.fill=fill(AZUL_SUB)
        cell.alignment=alin(h=h)
        cell.border=thin()
    ws.row_dimensions[r].height=16
    r += 1

    cat = c["cat"]
    bg_cat, fg_cat = cat_colors(cat)

    renglones = [
        ("(1)",  "Cartera Vigente",
                 "balanza",                  d["cartera_vigente"],  False, False),
        ("(2)",  "Cartera Vencida",
                 "balanza",                  d["cartera_vencida"],  False, False),
        ("(3)",  "Estimación preventiva para riesgos crediticios",
                 "balanza",                  d["estimacion_prev"],  False, True),
        ("(4)",  "Total de cartera de crédito neta  (1) + (2) − (3)",
                 "Art. 1 Bis 3",             c["cn"],               True,  False),
        ("(5)",  "Requerimientos de capitalización  (4) × 8%",
                 "Art. 1 Bis 3",             c["rq"],               True,  True),
        ("(6)",  "Capital Contable",
                 "balanza",                  c["cc"],               False, False),
        ("(7)",  "Certificados excedentes/voluntarios no elegibles",
                 "Art. 1 Bis 4 fr. II",      d["cert_vol"] or 0,    False, False),
        ("(8)",  "Financiamiento para adquisición de partes sociales",
                 "Art. 1 Bis 4 fr. III",     0,                     False, True),
        ("(9)",  "Capital neto  (6) − (7) − (8)",
                 "Art. 1 Bis 4",             c["kn"],               True,  False),
        ("(10)", "Nivel de capitalización  [(9) / (5)] × 100",
                 "Art. 1 Bis 6",             c["nc"],               True,  True),
    ]

    for num, concepto, fund, val, bold, sep in renglones:
        es_nivel = (num == "(10)")
        fh = GRIS_FILA if bold and not es_nivel else (bg_cat if es_nivel else BLANCO)
        fg = fg_cat if es_nivel else NEGRO
        brd = thick_bottom() if sep else bottom_only("00CCCCCC")

        ws.cell(row=r,column=1,value=num).font=font(size=9,color="00666666")
        ws.cell(row=r,column=1).alignment=alin(h="center")
        ws.cell(row=r,column=1).fill=fill(fh)
        ws.cell(row=r,column=1).border=brd

        c2=ws.cell(row=r,column=2,value=concepto)
        c2.font=font(bold=bold,size=9,color=fg); c2.fill=fill(fh)
        c2.alignment=alin(); c2.border=brd

        c3=ws.cell(row=r,column=3,value=fund)
        c3.font=font(size=8,color="00666666"); c3.fill=fill(fh)
        c3.alignment=alin(h="center"); c3.border=brd

        c4=ws.cell(row=r,column=4,value=val)
        c4.font=font(bold=bold,size=9,color=fg); c4.fill=fill(fh)
        c4.alignment=alin(h="right"); c4.border=brd
        c4.number_format = FMT_PCT if es_nivel else FMT_PESO_N
        if es_nivel and val is not None:
            c4.value = val / 100  # Excel guarda fracciones para porcentajes

        ws.row_dimensions[r].height=15
        r += 1

    # Resultado destacado
    r += 1
    merge_write(ws, r, 1, r, 2,
                f"Categoría de Capitalización: {cat}",
                bold=True, size=14, color=fg_cat, fill_hex=bg_cat, h="left")
    ws.cell(row=r,column=1).alignment=alin(h="left")
    c4=ws.cell(row=r,column=4,value=(c["nc"]/100 if c["nc"] else None))
    c4.font=font(bold=True,size=14,color=fg_cat); c4.fill=fill(bg_cat)
    c4.number_format=FMT_PCT; c4.alignment=alin(h="right")
    ws.row_dimensions[r].height=24
    r += 2

    # Obligaciones
    obligaciones = {
        "A": ["Notificar la clasificación en la asamblea inmediata siguiente."],
        "B": ["Notificar la clasificación en la asamblea inmediata siguiente."],
        "C": ["Adoptar medidas correctivas inmediatas.",
              "Notificar a la Asamblea en máximo 30 días (Art. 15, fracc. II).",
              "ALERTA: dos C consecutivas derivan en categoría D (Art. 15, fracc. III)."],
        "D": ["Abstenerse de operaciones de captación (Art. 15, fracc. IV).",
              "Iniciar disolución y liquidación.",
              "Notificar a la Asamblea en máximo 30 días (Art. 15, fracc. II)."],
    }.get(cat, [])

    cell=ws.cell(row=r,column=1,value="Obligaciones derivadas (Art. 15 LRASCAP):")
    cell.font=font(bold=True,size=9); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
    r += 1
    for ob in obligaciones:
        cell=ws.cell(row=r,column=1,value=f"• {ob}")
        cell.font=font(size=9); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
        r += 1

    r += 1
    nota=("La formulación y presentación de los estados financieros básicos es responsabilidad del "
          "Consejo de Administración (Art. 1 Bis 1). El cómputo rige salvo que el Comité de "
          "Supervisión Auxiliar obtenga uno distinto (Art. 1 Bis 6), en cuyo caso el del Comité "
          "será el definitivo. Contraparte: CSA / FOCOOP (no CNBV directamente).")
    cell=ws.cell(row=r,column=1,value=nota)
    cell.font=font(size=7.5,italic=True,color="00666666")
    cell.alignment=alin(wrap=True)
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
    ws.row_dimensions[r].height=36

# ── Función principal ──────────────────────────────────────────────────────────
def generar_excel(d, c, nombre, fecha_corte, operador=""):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja vacía por defecto

    hoja_balance(wb, d, c, nombre, fecha_corte)
    hoja_er(wb, d, c, nombre, fecha_corte)
    hoja_cap(wb, d, c, nombre, fecha_corte)

    # Propiedades del workbook
    wb.properties.title   = "Reporte Regulatorio Nivel Básico"
    wb.properties.creator = operador or "ACR Normativa CNBV"
    wb.properties.subject = "Anexo U — SOCAP Nivel de Operaciones Básico"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

