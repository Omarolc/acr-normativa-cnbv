"""
ACR Normativa CNBV — Dashboard Web v2
SOCAP Nivel de Operaciones Básico — con sistema de acceso por clave
"""
import os, io, textwrap, hashlib, hmac, json, base64
from datetime import date, datetime
from flask import Flask, request, jsonify, send_file, render_template, session
from flask_cors import CORS
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import cm
from reportlab.lib.colors import Color
from reportlab.pdfgen import canvas
from reportlab.lib import colors

app = Flask(__name__)
app.secret_key = os.environ.get("ACR_SESSION_KEY", "acr-session-dev-2026")
CORS(app)
W, H = letter

# ── Configuración de claves ───────────────────────────────────────────────────
MASTER_SECRET  = os.environ.get("ACR_LICENSE_SECRET", "")
MASTER_KEY     = os.environ.get("ACR_MASTER_KEY", "")  # sin default — debe configurarse en Railway
REVOCADAS      = set(os.environ.get("ACR_REVOKED_IDS","").upper().split(",")) - {""}

# ── Verificación de licencia ──────────────────────────────────────────────────
def _decodificar_clave(clave: str) -> dict:
    partes = clave.strip().rsplit(".", 1)
    if len(partes) != 2:
        raise ValueError("Formato inválido")
    payload_b64, firma_recibida = partes
    padding = 4 - len(payload_b64) % 4
    payload = json.loads(base64.urlsafe_b64decode(payload_b64 + "="*(padding%4)).decode())
    firma_esp = hmac.new(MASTER_SECRET.encode(), payload_b64.encode(), hashlib.sha256).hexdigest()[:16].upper()
    if not hmac.compare_digest(firma_recibida.upper(), firma_esp):
        raise ValueError("Clave inválida")
    return payload

def verificar_acceso(clave: str) -> dict:
    """
    Retorna dict con info del acceso:
      tipo: 'master' | 'prueba'
      titular, usos_restantes, expira, id
    """
    clave = clave.strip()

    # Clave maestra — solo si está configurada y coincide exactamente
    if MASTER_KEY and clave == MASTER_KEY:
        return {"tipo": "master", "titular": "Omar León Corona", "usos_restantes": 999}

    # Clave de prueba
    if not MASTER_SECRET:
        raise ValueError("Servidor sin configuración de licencias.")
    payload = _decodificar_clave(clave)
    lid = payload["id"].upper()
    if lid in REVOCADAS:
        raise PermissionError("Esta clave ha sido revocada. Contacta a Omar León Corona.")
    expira = date.fromisoformat(payload["expira"])
    if date.today() > expira:
        raise PermissionError(f"Esta clave expiró el {expira.strftime('%d/%m/%Y')}.")

    # Contar usos
    from licencias import _init_db, _contar_usos
    con = _init_db()
    usos = _contar_usos(con, lid)
    con.close()
    restantes = payload["max_usos"] - usos
    if restantes <= 0:
        raise PermissionError(f"Esta clave ya consumió sus {payload['max_usos']} reportes. Contacta a Omar León Corona.")

    return {
        "tipo": "prueba",
        "titular": payload["titular"],
        "id": lid,
        "max_usos": payload["max_usos"],
        "usos_restantes": restantes,
        "expira": str(expira),
    }

def generar_clave_prueba(titular: str, max_usos: int, expira: date) -> tuple:
    import secrets as _sec
    # ID legible: iniciales del titular + folio hex corto (ej. MM-3F9A1C)
    palabras = titular.split("—")[0].strip().split()
    iniciales = "".join(p[0].upper() for p in palabras if p)[:3]
    folio = _sec.token_hex(3).upper()
    clave_id = f"{iniciales}-{folio}"
    payload = {
        "titular": titular,
        "max_usos": max_usos,
        "expira": str(expira),
        "id": clave_id
    }
    b64 = base64.urlsafe_b64encode(
        json.dumps(payload, ensure_ascii=False).encode()
    ).decode().rstrip("=")
    firma = hmac.new(MASTER_SECRET.encode(), b64.encode(), hashlib.sha256).hexdigest()[:16].upper()
    return f"{b64}.{firma}", clave_id

# ── Extracción de datos ───────────────────────────────────────────────────────
def norm(s):
    import unicodedata
    return unicodedata.normalize("NFD", str(s or "").lower()).encode("ascii","ignore").decode().strip()

CLAVES = {
    "efectivo":       ["otras disponibilidades","total disponibilidades","caja","bancos","efectivo","disponibilidades"],
    "cartera_vigente":["total cartera de cred. vigente","total cartera de credito vigente","total cartera vigente"],
    "cartera_vencida":["total cartera de cred. vencida","total cartera de credito vencida","total cartera vencida"],
    "estimacion_prev":["est. prev. p/ riesgo","estimacion preventiva para riesgo","estimacion preventiva"],
    "otras_cxc":      ["otras cuentas por cobrar"],
    "bienes_adj":     ["bienes adjudicados"],
    "inmuebles":      ["inmuebles, mobiliario","mobiliario y equipo"],
    "otros_activos":  ["otros activos","cargos diferidos y pagos anticipados"],
    "dep_exig_inm":   ["depositos de exigibilidad inmediata"],
    "dep_plazo":      ["depositos a plazo","deposito a plazo"],
    "cuentas_sin_mov":["cuentas sin movimiento"],
    "prestamos_cp":   ["de corto plazo"],
    "prestamos_lp":   ["de largo plazo"],
    "otras_cxp":      ["acreedores diversos","otras cuentas por pagar"],
    "cert_ord":       ["capital social total","capital social"],
    "cert_vol":       ["aportacion voluntaria","certificados excedentes","aportaciones voluntarias"],
    "reservas":       ["fondo de reserva","reservas de capital","reserva"],
    "result_ant":     ["resultado de ejercicios anteriores"],
    "result_neto_bg": ["resultado neto"],
    "ingresos_int":   ["ingresos por intereses"],
    "gastos_int":     ["gastos por intereses"],
    "est_riesgos_er": ["estimacion preventiva para riesgos","est. prev. p/ riesgo"],
    "otros_ing":      ["otros productos","resultado por intermediacion","comisiones y tarifas cobradas"],
    "gastos_adm":     ["gastos de administracion y promocion","gastos de administracion"],
}
ER_SET = {"ingresos_int","gastos_int","est_riesgos_er","otros_ing","gastos_adm"}

def buscar(rows, claves, absoluto=False):
    for row in rows:
        textos = [norm(c) for c in row if isinstance(c, str) and len(str(c).strip()) > 2]
        nums   = [c for c in row if isinstance(c, (int, float)) and abs(c) > 0.01]
        if not nums or not textos: continue
        for clave in claves:
            nc = norm(clave)
            if any(nc in t or (len(nc) > 12 and t.startswith(nc[:12])) for t in textos):
                return abs(max(nums, key=abs)) if absoluto else max(nums, key=abs)
    return 0.0

def extraer(wb):
    nms = [n.upper() for n in wb.sheetnames]
    def hoja(pref):
        for i, n in enumerate(nms):
            if any(p in n for p in pref): return wb.worksheets[i]
        return wb.worksheets[0]
    ws_bg = hoja(["BG","BALANCE","SITUACION","ACTIVO"])
    ws_er = hoja(["ED","ER","RESULTADO","RESULT"])
    rows_bg = [[c.value for c in r] for r in ws_bg.iter_rows()]
    rows_er = [[c.value for c in r] for r in ws_er.iter_rows()]
    d = {}
    for rubro, claves in CLAVES.items():
        rows = rows_er if rubro in ER_SET else rows_bg
        d[rubro] = buscar(rows, claves, absoluto=(rubro == "estimacion_prev"))
    nombre, fecha = "", ""
    for row in rows_bg[:10]:
        for cell in row:
            if not isinstance(cell, str): continue
            nu = norm(cell)
            if len(cell) > 15 and ("cooperativa" in nu or "s.c." in nu or "sc de" in nu):
                nombre = cell.strip()
            if ("al " in nu or "estado de situacion" in nu) and "20" in cell:
                fecha = cell.strip()
    return d, nombre, fecha

def calcular(d):
    cn  = d["cartera_vigente"] + d["cartera_vencida"] - d["estimacion_prev"]
    rq  = cn * 0.08
    cc  = d["cert_ord"] + d["cert_vol"] + d["reservas"] + d["result_ant"] + d["result_neto_bg"]
    kn  = cc
    nc  = (kn / rq * 100) if rq > 0 else None
    cat = ("A" if nc and nc >= 150 else "B" if nc and nc >= 100 else
           "C" if nc and nc >= 50  else "D" if nc is not None else "—")
    ta  = d["efectivo"] + cn + d["otras_cxc"] + d["bienes_adj"] + d["inmuebles"] + d["otros_activos"]
    tp  = d["dep_exig_inm"]+d["dep_plazo"]+d["cuentas_sin_mov"]+d["prestamos_cp"]+d["prestamos_lp"]+d["otras_cxp"]
    rf  = d["ingresos_int"] - d["gastos_int"]
    rfaj= rf - d["est_riesgos_er"]
    rn  = rfaj + d["otros_ing"] - d["gastos_adm"]
    return dict(cn=cn,rq=rq,cc=cc,kn=kn,nc=nc,cat=cat,
                ta=ta,tp=tp,cuadra=abs(ta-(tp+cc))<500,rf=rf,rfaj=rfaj,rn=rn)

# ── PDF ───────────────────────────────────────────────────────────────────────
CAT_COLOR = {"A":colors.HexColor("#1a7a1a"),"B":colors.HexColor("#7a5a00"),
             "C":colors.HexColor("#a04000"), "D":colors.HexColor("#b00000")}

def p(n):
    if n is None: return "—"
    return f"${abs(n):>14,.2f}" if n >= 0 else f"(${abs(n):>13,.2f})"

def generar_pdf(d, c, nombre, fecha_corte, operador, es_prueba=False, titular=""):
    buf = io.BytesIO()
    cv  = canvas.Canvas(buf, pagesize=letter)
    cv.setAuthor("Omar León Corona — ACR Normativa CNBV")

    def marca_agua():
        if not es_prueba: return
        cv.saveState()
        cv.setFillColor(Color(0.65,0.1,0.1,alpha=0.08))
        cv.setFont("Helvetica-Bold", 55)
        cv.translate(W/2, H/2)
        cv.rotate(42)
        cv.drawCentredString(0, 50, "PRUEBA")
        cv.drawCentredString(0, -70, "PRUEBA")
        cv.restoreState()

    def encabezado(titulo, subtitulo):
        cv.setLineWidth(1.5)
        cv.line(1.5*cm, H-1.8*cm, W-1.5*cm, H-1.8*cm)
        marca_agua()
        cv.setFont("Helvetica-Bold", 9)
        cv.drawCentredString(W/2, H-2.5*cm, nombre or "SOCIEDAD COOPERATIVA DE AHORRO Y PRÉSTAMO")
        cv.setFont("Helvetica-Bold", 10.5)
        cv.drawCentredString(W/2, H-3.3*cm, titulo)
        cv.setFont("Helvetica", 8.5)
        cv.drawCentredString(W/2, H-3.9*cm, subtitulo)
        cv.setFont("Helvetica", 7.5)
        cv.drawCentredString(W/2, H-4.4*cm, "(Cifras en pesos)")
        cv.setLineWidth(0.4)
        cv.line(1.5*cm, H-4.8*cm, W-1.5*cm, H-4.8*cm)

    def pie():
        cv.setLineWidth(0.4)
        cv.line(1.5*cm, 1.6*cm, W-1.5*cm, 1.6*cm)
        cv.setFont("Helvetica", 7)
        cv.drawString(1.5*cm, 1.2*cm,
            f"Art. 1 Bis 1 Disp. CNBV · Contraparte: CSA (FOCOOP) · "
            f"Desarrollado por Omar León Corona © 2026 · Elaborado: {operador}")
        cv.drawRightString(W-1.5*cm, 1.2*cm, f"Corte: {fecha_corte}")
        if es_prueba:
            cv.saveState()
            cv.setFont("Helvetica-Bold", 6)
            cv.setFillColor(Color(0.6,0.1,0.1))
            cv.drawCentredString(W/2, 0.6*cm,
                f"REPORTE DE PRUEBA — Autorizado para: {titular} — No válido para entrega oficial al CSA")
            cv.restoreState()

    LX,RX,NL,NR = 1.8,10.8,10.3,19.3
    def fila(y,tl,il,tr,ir,bl=False,br=False):
        cv.setFont("Helvetica-Bold" if bl else "Helvetica",8)
        if tl: cv.drawString(LX*cm,y,tl)
        if il: cv.setFont("Helvetica",8); cv.drawRightString(NL*cm,y,il)
        cv.setFont("Helvetica-Bold" if br else "Helvetica",8)
        if tr: cv.drawString(RX*cm,y,tr)
        if ir: cv.setFont("Helvetica",8); cv.drawRightString(NR*cm,y,ir)
    def sl(y): cv.line(LX*cm,y-1,NL*cm,y-1)
    def sr(y): cv.line(RX*cm,y-1,NR*cm,y-1)

    # Página 1 — Balance
    encabezado(f"BALANCE GENERAL AL {fecha_corte.upper()}","")
    cv.setLineWidth(0.35)
    y=H-5.3*cm
    cv.setFont("Helvetica-Bold",8.5)
    cv.drawString(LX*cm,y,"ACTIVO"); cv.drawString(RX*cm,y,"PASIVO Y CAPITAL"); y-=.5*cm
    cn=c["cn"]
    fila(y,"EFECTIVO",p(d["efectivo"]),"DEPÓSITOS","",True,True); y-=.45*cm
    fila(y,"","","  Exigibilidad inmediata",p(d["dep_exig_inm"])); y-=.45*cm
    fila(y,"CARTERA VIGENTE",p(d["cartera_vigente"]),"  A plazo",p(d["dep_plazo"]),True); y-=.45*cm
    fila(y,"CARTERA VENCIDA",p(d["cartera_vencida"]),"  Sin movimiento",p(d["cuentas_sin_mov"]),True); y-=.45*cm
    sl(y); fila(y,"TOTAL CARTERA",p(d["cartera_vigente"]+d["cartera_vencida"]),"","",True); sr(y); y-=.45*cm
    fila(y,"(−) ESTIMACIÓN PREVENTIVA",p(d["estimacion_prev"]),"PRÉSTAMOS BANCARIOS","",True,True); y-=.45*cm
    sl(y); fila(y,"CARTERA NETA",p(cn),"  Corto plazo",p(d["prestamos_cp"]),True); y-=.45*cm
    fila(y,"OTRAS CxC (NETO)",p(d["otras_cxc"]),"  Largo plazo",p(d["prestamos_lp"])); y-=.45*cm
    sr(y); fila(y,"BIENES ADJUDICADOS",p(d["bienes_adj"]),"OTRAS CxP",p(d["otras_cxp"]),False,True); y-=.45*cm
    sr(y); fila(y,"INMUEBLES Y EQUIPO (NETO)",p(d["inmuebles"]),"TOTAL PASIVO",p(c["tp"]),False,True); y-=.45*cm
    fila(y,"OTROS ACTIVOS",p(d["otros_activos"]),"CAPITAL CONTRIBUIDO","",False,True); y-=.4*cm
    fila(y,"","","  Capital social / Cert. ordinarios",p(d["cert_ord"])); y-=.4*cm
    fila(y,"","","  Certificados excedentes/voluntarios",p(d["cert_vol"])); y-=.4*cm
    sr(y); fila(y,"","","CAPITAL GANADO","",False,True); y-=.4*cm
    fila(y,"","","  Reservas de capital",p(d["reservas"])); y-=.4*cm
    fila(y,"","","  Resultado ejercicios anteriores",p(d["result_ant"])); y-=.4*cm
    fila(y,"","","  Resultado neto del periodo",p(d["result_neto_bg"])); y-=.45*cm
    sr(y); fila(y,"","","TOTAL CAPITAL CONTABLE",p(c["cc"]),False,True); y-=.55*cm
    cv.setLineWidth(1.2); sl(y); sr(y); cv.setLineWidth(0.35)
    fila(y,"TOTAL ACTIVO",p(c["ta"]),"TOTAL PASIVO Y CAPITAL",p(c["tp"]+c["cc"]),True,True)
    y-=1.8*cm; cv.setLineWidth(0.5)
    cv.line(2*cm,y,9*cm,y); cv.line(12*cm,y,19.5*cm,y)
    cv.setFont("Helvetica",7.5)
    cv.drawCentredString(5.5*cm,y-.4*cm,"Presidente del Consejo de Administración")
    cv.drawCentredString(15.75*cm,y-.4*cm,"Director o Gerente General")
    cv.setFont("Helvetica",6.5)
    cv.drawCentredString(5.5*cm,y-.75*cm,"(entrega semestral impresa)")
    cv.drawCentredString(15.75*cm,y-.75*cm,"(entrega semestral impresa)")
    pie()

    # Página 2 — Estado de Resultados
    cv.showPage()
    encabezado("ESTADO DE RESULTADOS",f"DEL 1 DE ENERO AL {fecha_corte.upper()}")
    cv.setLineWidth(0.35); y=H-5.5*cm; LXe,NUMe=3.0,17.0
    def fer(y,con,val,bold=False,doble=False,sep=False):
        if sep:
            cv.setLineWidth(1.2 if doble else 0.4)
            cv.line(LXe*cm,y+.35*cm,NUMe*cm,y+.35*cm); cv.setLineWidth(0.35)
        cv.setFont("Helvetica-Bold" if bold else "Helvetica",9)
        cv.drawString(LXe*cm,y,con)
        if val is not None:
            cv.setFont("Helvetica",9); cv.drawRightString(NUMe*cm,y,p(val))
    fer(y,"Ingresos por intereses",d["ingresos_int"]); y-=.55*cm
    fer(y,"Gastos por intereses",d["gastos_int"]); y-=.55*cm
    fer(y,"RESULTADO FINANCIERO",c["rf"],bold=True,sep=True); y-=.6*cm
    fer(y,"Estimación preventiva para riesgos crediticios",d["est_riesgos_er"]); y-=.55*cm
    fer(y,"RESULTADO FINANCIERO AJUSTADO POR RIESGOS CREDITICIOS",c["rfaj"],bold=True,sep=True); y-=.6*cm
    fer(y,"Otros ingresos (egresos) de la operación",d["otros_ing"]); y-=.5*cm
    fer(y,"Gastos de administración y promoción",d["gastos_adm"]); y-=.55*cm
    fer(y,"RESULTADO NETO",c["rn"],bold=True,sep=True,doble=True)
    y-=2.5*cm; cv.setLineWidth(0.5)
    cv.line(2*cm,y,9*cm,y); cv.line(12*cm,y,19.5*cm,y)
    cv.setFont("Helvetica",7.5)
    cv.drawCentredString(5.5*cm,y-.4*cm,"Presidente del Consejo de Administración")
    cv.drawCentredString(15.75*cm,y-.4*cm,"Director o Gerente General")
    cv.setFont("Helvetica",6.5)
    cv.drawCentredString(5.5*cm,y-.75*cm,"(entrega semestral impresa)")
    cv.drawCentredString(15.75*cm,y-.75*cm,"(entrega semestral impresa)")
    pie()

    # Página 3 — Cómputo
    cv.showPage()
    encabezado("CÓMPUTO DEL NIVEL DE CAPITALIZACIÓN",
               f"CIFRAS AL {fecha_corte.upper()} · Arts. 1 Bis 3, 1 Bis 4 y 1 Bis 6")
    cv.setLineWidth(0.35); y=H-5.5*cm
    LXc,NCL,NCR=1.3,15.5,20.3
    cv.setFont("Helvetica-Bold",7.5)
    cv.drawString(LXc*cm,y,"#"); cv.drawString((LXc+0.7)*cm,y,"Concepto")
    cv.drawRightString(NCL*cm,y,"Fundamento"); cv.drawRightString(NCR*cm,y,"Importe")
    y-=.3*cm; cv.setLineWidth(0.8); cv.line(LXc*cm,y,NCR*cm,y); cv.setLineWidth(0.35); y-=.5*cm
    rengs=[
        ("(1)","Cartera Vigente",d["cartera_vigente"],"balanza",False,False),
        ("(2)","Cartera Vencida",d["cartera_vencida"],"balanza",False,False),
        ("(3)","Estimación preventiva para riesgos crediticios",d["estimacion_prev"],"balanza",False,False),
        ("(4)","Total de cartera de crédito neta  (1)+(2)−(3)",c["cn"],"Art. 1 Bis 3",True,False),
        ("(5)","Requerimientos de capitalización  (4)×8%",c["rq"],"Art. 1 Bis 3",True,False),
        ("(6)","Capital Contable",c["cc"],"balanza",False,False),
        ("(7)","Certificados excedentes/voluntarios no elegibles",d["cert_vol"],"Art. 1 Bis 4 fr. II",False,False),
        ("(8)","Financiamiento para adquisición de partes sociales",0,"Art. 1 Bis 4 fr. III",False,False),
        ("(9)","Capital neto  (6)−(7)−(8)",c["kn"],"Art. 1 Bis 4",True,False),
        ("(10)","Nivel de capitalización  [(9)/(5)]×100",c["nc"],"Art. 1 Bis 6",True,True),
    ]
    for num,con,val,fund,bold,esNiv in rengs:
        cv.setFont("Helvetica",7); cv.drawString(LXc*cm,y,num)
        cv.setFont("Helvetica-Bold" if bold else "Helvetica",7)
        cv.drawString((LXc+0.7)*cm,y,con)
        cv.setFont("Helvetica",6.5); cv.drawRightString(NCL*cm,y,fund)
        cv.setFont("Helvetica-Bold" if bold else "Helvetica",7.5)
        txt=(f"{c['nc']:.2f}%" if c["nc"] is not None else "—") if esNiv else p(val)
        cv.drawRightString(NCR*cm,y,txt); y-=.52*cm
        if num in ("(3)","(5)","(8)"):
            cv.line((LXc+0.7)*cm,y+.38*cm,NCR*cm,y+.38*cm)
    y-=.2*cm
    cv.setLineWidth(1.2); cv.line(LXc*cm,y+.3*cm,NCR*cm,y+.3*cm)
    cv.setLineWidth(0.4); cv.line(LXc*cm,y+.15*cm,NCR*cm,y+.15*cm)
    cat=c["cat"]
    cv.setFillColor(CAT_COLOR.get(cat,colors.black)); cv.setFont("Helvetica-Bold",11)
    cv.drawString(LXc*cm,y-1*cm,f"Categoría de Capitalización:  {cat}")
    cv.setFont("Helvetica",9)
    cv.drawString(LXc*cm+9*cm,y-1*cm,f"Nivel: {c['nc']:.2f}%" if c['nc'] else "Nivel: —")
    cv.setFillColor(colors.black); y-=1.8*cm
    obs={"A":["Notificar en asamblea inmediata siguiente."],
         "B":["Notificar en asamblea inmediata siguiente."],
         "C":["Adoptar medidas correctivas inmediatas.",
              "Notificar a la Asamblea en máximo 30 días (Art. 15, fracc. II).",
              "ALERTA: dos C consecutivas derivan en D (Art. 15, fracc. III)."],
         "D":["Abstenerse de captación (Art. 15, fracc. IV).",
              "Iniciar disolución y liquidación.",
              "Notificar a la Asamblea en máximo 30 días (Art. 15, fracc. II)."]}.get(cat,[])
    cv.setFont("Helvetica-Bold",8); cv.drawString(LXc*cm,y,"Obligaciones derivadas (Art. 15 LRASCAP):"); y-=.4*cm
    cv.setFont("Helvetica",8)
    for ob in obs:
        cv.drawString((LXc+0.3)*cm,y,f"• {ob}"); y-=.4*cm
    y-=.4*cm
    nota=("La formulación y presentación es responsabilidad del Consejo de Administración (Art. 1 Bis 1). "
          "El cómputo rige salvo verificación del CSA (Art. 1 Bis 6). "
          "Sistema desarrollado por Omar León Corona © 2026 — ACR Normativa CNBV.")
    cv.setFont("Helvetica",7.5)
    for linea in textwrap.wrap(nota,118):
        cv.drawString(LXc*cm,y,linea); y-=.35*cm
    pie(); cv.save(); buf.seek(0)
    return buf

# ── Rutas ─────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/login", methods=["POST"])
def login():
    clave = (request.json or {}).get("clave","").strip()
    try:
        info = verificar_acceso(clave)
        return jsonify({"ok": True, **info})
    except (ValueError, PermissionError) as e:
        return jsonify({"ok": False, "error": str(e)}), 401

@app.route("/api/generar-clave", methods=["POST"])
def api_generar_clave():
    """Solo disponible para la clave maestra."""
    data     = request.json or {}
    clave_op = data.get("clave_operador","").strip()
    if clave_op != MASTER_KEY:
        return jsonify({"error": "No autorizado"}), 403
    titular  = data.get("titular","")
    max_usos = int(data.get("max_usos", 5))
    expira   = date.fromisoformat(data.get("expira", "2026-08-31"))
    clave, lid = generar_clave_prueba(titular, max_usos, expira)
    # Guardar en log
    try:
        log = os.path.join(os.path.dirname(__file__), "licencias_emitidas.jsonl")
        with open(log,"a",encoding="utf-8") as f:
            f.write(json.dumps({
                "timestamp": datetime.utcnow().isoformat(),
                "titular": titular, "id": lid,
                "max_usos": max_usos, "expira": str(expira), "clave": clave,
            }, ensure_ascii=False) + "\n")
    except Exception:
        pass
    return jsonify({"clave": clave, "id": lid, "titular": titular,
                    "max_usos": max_usos, "expira": str(expira)})

@app.route("/api/procesar", methods=["POST"])
def procesar():
    if "archivo" not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400
    f       = request.files["archivo"]
    fecha   = request.form.get("fecha","2026-06-30")
    operador= request.form.get("operador","")
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        d, nombre, fecha_raw = extraer(wb)
        c = calcular(d)
        return jsonify({"nombre":nombre,"fecha_raw":fecha_raw,
                        "datos":d,"calculo":c,"fecha_corte":fecha,"operador":operador})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def _validar_descarga(data):
    clave = data.get("clave_sesion","").strip()
    try:
        info = verificar_acceso(clave)
        es_prueba = info["tipo"] == "prueba"
        if es_prueba:
            # Registrar uso
            from licencias import _init_db, _registrar_uso
            con = _init_db()
            _registrar_uso(con, info["id"], info["titular"], request.remote_addr or "")
            con.close()
        return es_prueba, info.get("titular",""), None
    except (ValueError, PermissionError) as e:
        return None, None, ({"error": str(e)}, 403)

@app.route("/api/pdf", methods=["POST"])
def pdf():
    data = request.json
    es_prueba, titular, err = _validar_descarga(data)
    if err: return jsonify(err[0]), err[1]
    d=data["datos"]; c=data["calculo"]
    nombre=data.get("nombre",""); fecha=data.get("fecha_corte","2026-06-30")
    operador=data.get("operador","")
    buf = generar_pdf(d,c,nombre,fecha,operador,es_prueba=es_prueba,titular=titular)
    sfx = "_PRUEBA" if es_prueba else ""
    return send_file(buf,mimetype="application/pdf",as_attachment=True,
                     download_name=f"AnexoU_{nombre[:20].replace(' ','_')}_{fecha}{sfx}.pdf")

@app.route("/api/excel", methods=["POST"])
def excel():
    from generar_excel_anexou import generar_excel
    data = request.json
    es_prueba, titular, err = _validar_descarga(data)
    if err: return jsonify(err[0]), err[1]
    d=data["datos"]; c=data["calculo"]
    nombre=data.get("nombre",""); fecha=data.get("fecha_corte","2026-06-30")
    operador=data.get("operador","")
    buf = generar_excel(d,c,nombre,fecha,operador,es_prueba=es_prueba,titular_prueba=titular)
    sfx = "_PRUEBA" if es_prueba else ""
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=f"AnexoU_{(nombre or 'SOCAP')[:20].replace(' ','_')}_{fecha}{sfx}.xlsx")

if __name__ == "__main__":
    port = int(os.environ.get("PORT",5000))
    app.run(host="0.0.0.0",port=port,debug=False)
